import scipy.io
import numpy as np
import time
import os
import multiprocessing
from multiprocessing import Process, Queue
import math
import openpyxl

resolution_mat='40_row_matlab'
mat_folder_path=f'C:/Users/asdfg/Desktop/AMD_sunkyu/projects/InterpretationReID/{resolution_mat}'
mat_filelist=os.listdir(mat_folder_path)
gal_label_mat= scipy.io.loadmat('C:/Users/asdfg/Desktop/AMD_sunkyu/projects/InterpretationReID/gallery_label.mat')
person_color_mat=scipy.io.loadmat('C:/Users/asdfg/Desktop/AMD_sunkyu/projects/InterpretationReID/person_color.mat')
gal_label=gal_label_mat['gallery_label']

gal_images=os.listdir('C:/docker_volume/AMD/datasets/Market1501_/bounding_box_test')

up_color_dict={18:"black",19:"blue",20:"gray",21:"green",22:"purple",23:"red",24:"white",25:"yellow",0:"unknown"} #상의색 8개
down_color_dict={4:"black",5:"blue",6:"brown",7:"gray",8:"green",9:"pink",10:"purple",11:"white",12:"yellow",0:"unknown"} #하의색 9개


# fake_result 엑셀 파일 불러오기 (5개 색 이미지 넣어서 갤러리 라베링한 값)
excel_file_path = 'fake_result.xlsx'
sheet_name='Total'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook[sheet_name]

gal_pre_up_list = [cell.value for cell in sheet[1]]
gal_pre_down_list = [cell.value for cell in sheet[2]]



#이미지 id와 750명 인덱스 맞추기 
a=gal_label_mat['gallery_label'][0].tolist()
list_p=[]

for i in a:
    list_p.append(i)
list_p=set(list_p)
list_p=list(list_p)
del list_p[0]

image_id={}
for i in range(750):
    image_id[list_p[i]]=i
#print(image_id)




def work(id, mat_list, result,up_rank,down_rank,folder_path):
    person_compare=1 # 이전 사람과 달라지는지 비교하기 위한 변수
    per_count=0 #

    count_list=[]
    id_list=[]
    TP_list=[]
    FP_list=[]
    TN_list=[]
    FN_list=[]
    F1_list=[]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title='Total'
    sheet2 = workbook.create_sheet(title="Filtered_index")
    sheet3 = workbook.create_sheet(title="Filtered_dist")
    sheet4 = workbook.create_sheet(title="Filtered_filename")
    res=['per_id','image_num','id_list','TP_list','FP_list','TN_list','FN_list','F1_list']
    sheet.append(res)

    rank1_list=[]
    rank5_list=[]
    rank10_list=[]
    
    for try_num,mat_file in enumerate(mat_list):
        load_data = scipy.io.loadmat(mat_file)

        dist_total=load_data['dist_total'] #쿼리와 갤러리 간 전체 거리값 
        dist_att=load_data['dist_att'] #dist_fake_stack belong to 0.0~1.0, 쿼리와 갤러리 간 속성별 거리값

        person_num=int(mat_file.split('/')[-1].split('_')[0])

        per_count=image_id[person_num]

        query_up=up_color_dict[person_color_mat['upcolor'][0][per_count]] #상의색 구하기
        query_down=down_color_dict[person_color_mat['downcolor'][0][per_count]] # 하의색 구하기

        #print(query_up,query_down)

        count=0
        count_gal=0
        gal_num=[]

        gal_label_dict={} # 신원별 이미지 리스트가 들어있음. 
        for index, value in enumerate(gal_label.tolist()[0]):
            if value in gal_label_dict:
                gal_label_dict[value].append(index)
            else:
                gal_label_dict[value]=[index]

        #confusion matrix
        TP = FP = TN = FN = 0

        #max가맞음
        gal_count=-1
        person_gal_id=0

        for index,i in enumerate(dist_att):
            gal_list=i #첫번째 갤러리에 대한 속성별 거리 

            up_color_att={}
            down_color_att={}
            up_color_val=gal_list[18:26] # 상의색 속성 거리 저장
            down_color_val=gal_list[4:13] # 하의색 속성 거리 저장

            #속성 값 중 상위 세개 값 담기
            up_rank3_value=sorted(up_color_val,reverse=True)[:3]
            down_rank3_value=sorted(down_color_val,reverse=True)[:3]

            #갤러리 속성별 거리 중 상,하의 값:인덱스 딕셔너리 생성
            for att_idx, y in enumerate(gal_list):
                if(18<=att_idx and att_idx<26):
                    up_color_att[y]=att_idx
                if(4<=att_idx and att_idx<13):
                    down_color_att[y]=att_idx
            
            gal_fake_up_idx=[]
            gal_fake_down_idx=[]

            #상, 하의 값 중 rank3 값 저장 
            for y in range(len(up_rank3_value)):
                gal_fake_up_idx.append(up_color_att[up_rank3_value[y]])
                gal_fake_down_idx.append(down_color_att[down_rank3_value[y]])
            
            gal_fake_up=[]
            gal_fake_down=[]

            #상, 하의 값 중 rank3 별 색상 속성값 대입
            for y in range(len(gal_fake_up_idx)):
                gal_fake_up.append(up_color_dict[gal_fake_up_idx[y]])
                gal_fake_down.append(down_color_dict[gal_fake_down_idx[y]])

            # 실제 갤러리의 상, 하의 색상 대입
            for x in gal_label_dict: #x 최댓값 = 1501
                if(index in gal_label_dict[x]):
                    if(person_gal_id!=x): #다음 인물로 넘어간다면?
                        person_gal_id=x # 다음 인물로 넘겨주고
                        gal_count+=1 # 갤러리 카운트 +1
                    if(person_gal_id!=0):
                        gal_real_up=up_color_dict[person_color_mat['upcolor'][0][gal_count]]
                        gal_real_down=down_color_dict[person_color_mat['downcolor'][0][gal_count]] 
                        break
                    else: #0번 인물들은 상, 하의 색 unknown으로 평가 
                        gal_real_up='unknown'
                        gal_real_down='unknown'
                        break
            #필터링 구간 
            if(query_up=='unknown'): #일단 실험 
                query_up=gal_fake_up[0]
            if(query_down=='unknown'):
                query_down=gal_fake_down[0]

            if((query_up == gal_pre_up_list[count_gal]) and (query_down == gal_pre_down_list[count_gal])): #상, 하의 정답 비교해서 완전히 일치하는 갤러리 인덱스 확인     #필터링 안한것
                gal_num.append(count_gal) # 잔여 갤러리 목록 추가 
                count+=1#필터링된 갤러리 이미지 개수를 세기위한 카운트
                if(query_up==gal_real_up and query_down==gal_real_down): #실제 
                    TN+=1
                else:
                    FP+=1
            else: # 필터링 했다. 
                if(query_up==gal_real_up and query_down==gal_real_down): #
                    FN+=1
                else:
                    TP+=1
            count_gal+=1 #갤러리에서 몇번째 갤러리 인물인지 세기 위함
            
        count_person=0
        gal_filter_dict={} #신원별 남아있는 갤러리 수 확인을 위한 딕셔너리
        for i in gal_label_dict: #i는 인물넘버를 의미함
            for j in gal_num:
                if(j in gal_label_dict[i]):
                    count_person+=1
                    gal_filter_dict[i]=count_person
            count_person=0

        # 20231119 2200v
        # if(TP + FP) == 0:
        #     print("ERRR")
        #     break
        # if(TP + FN) == 0:
        #     print("ERRR")
        #     break

        # percision = TP / (TP + FP)
        # recall = TP / (TP + FN)

        # if(percision == 0):
        #     print("ERRR")
        #     break
        # if(recall == 0):
        #     print("ERRR")
        #     break


        #F1_score = 2 * percision * recall / (percision + recall)
        F1_score=1
        sheet2.append(gal_num) # 인물 인덱스 번호 추가 

        ### Rank 계산하기 ###

        #필터링된 갤러리 이미지의 거리, 파일 명 추가
        filtered_dist_list=[] # 필터링된 갤러리 이미지 거리
        filtered_filename_list=[] #필터링된 갤러리 이미지 파일네임 
        for i in gal_num:
            filtered_dist_list.append(dist_total[0][i])
            filtered_filename_list.append(gal_images[i])
        sheet3.append(filtered_dist_list)
        sheet4.append(filtered_filename_list) 

        dict_fgal={}

        gal_rank10=sorted(filtered_dist_list)[:10] #필터링된 갤러리중 상위 10개 거리값 

        #거리값별 신원을 갖고있는 딕셔너리 생성 -> {거리값:신원}
        for i in range(len(filtered_dist_list)):
            dict_fgal[filtered_dist_list[i]]=int(filtered_filename_list[i].split('_')[0])
        
        gal_rank10_list=[]
        for i in gal_rank10:
            gal_rank10_list.append(dict_fgal[i])
        
        rank1=rank5=rank10=0
        if(person_num in gal_rank10_list[:1]):
            rank1+=1
            rank1_list.append(rank1)
        if(person_num in gal_rank10_list[:5]):
            rank5+=1
            rank5_list.append(rank5)
        if(person_num in gal_rank10_list[:10]):
            rank10+=1
            rank10_list.append(rank10)

        sheet.append([person_num,count,len(gal_filter_dict),TP,FP,TN,FN,F1_score])

        count_list.append(count)
        id_list.append(len(gal_filter_dict))
        TP_list.append(TP)
        FP_list.append(FP)
        TN_list.append(TN)
        FN_list.append(FN)
        F1_list.append(F1_score)
        
        if(try_num%50==0):
            print(f'{id}스레드 {try_num}작업 쿼리({mat_file}) 결과 -')

    # for g in range(len(count_list)):
    #     sheet.append([count_list[g],id_list[g],TP_list[g],FP_list[g],TN_list[g],FN_list[g],F1_list[g]])

    ret = {}
    ret['count_list'] = count_list
    ret['id_list'] = id_list
    ret['TP_list'] = TP_list
    ret['FP_list'] = FP_list
    ret['TN_list'] = TN_list
    ret['FN_list'] = FN_list
    ret['F1_list'] = F1_list
    ret['rank1_list']=rank1_list
    ret['rank5_list']=rank5_list
    ret['rank10_list']=rank10_list

    workbook.save(f'{folder_path}/{id}thread_.xlsx')
    workbook.close()

    result.put(ret)
    result.put('STOP')

    print(f"{id} Done")

if __name__ == "__main__":
    start_time=time.time()

    mat_paths = []
    for mat_file in mat_filelist:
        mat_paths.append(f'{mat_folder_path}/{mat_file}')
    num_matfiles = len(mat_paths) 

    ### 쓰레드 수
    num_core = 14

    #상의 하의 랭크 
    rank_up=1
    rank_down=1   
    ####

    #폴더 생성
    folder_path=f'{resolution_mat.split("_")[0]}_1122_test_row_rank_{rank_up}_{rank_down}'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    step = math.trunc(num_matfiles/num_core)

    ret_list = []
    th_list = []
    for i in range(num_core):
        START = step * i
        
        END = START + step
        if i == num_core-1:
            END = END + num_matfiles % num_core

        print(START, END, num_matfiles)
        result = Queue()
        
        th = Process(target=work, args=(i, mat_paths[START:END], result,rank_up,rank_down,folder_path))

        th_list.append(th)
        ret_list.append(result)
        th.start()

    for th in th_list:
        th.join()
        

    print("Join Done")

    count_list=[]
    id_list=[]
    TP_list=[]
    FP_list=[]
    TN_list=[]
    FN_list=[]
    F1_list=[]

    rank1_list=[]
    rank5_list=[]
    rank10_list=[]

    for result in ret_list: 
        while True:
            tmp = result.get()
            if tmp == 'STOP':
                break
            else:
                ret = tmp      
                count_list.extend(ret['count_list'])
                id_list.extend(ret['id_list'])
                TP_list.extend(ret['TP_list'])
                FP_list.extend(ret['FP_list'])
                TN_list.extend(ret['TN_list'])
                FN_list.extend(ret['FN_list'])
                F1_list.extend(ret['F1_list'])
                rank1_list.extend(ret['rank1_list'])
                rank5_list.extend(ret['rank5_list'])
                rank10_list.extend(ret['rank10_list'])

    end_time=time.time()
    elapsed_time = end_time - start_time

    with open(f'{resolution_mat.split("_")[0]}_test_row_rank_{rank_up}_{rank_down}.txt','w') as file:
        file.writelines(f"필터링된 갤러리 이미지 개수: {sum(count_list)/len(count_list)}")
        file.writelines(f"필터링된 잔여 신원 개수: {sum(id_list)/len(id_list)}") #신원 개수 구하기
        file.writelines(f"True Positive (TP):{sum(TP_list)/len(TP_list)}")
        file.writelines(f"False Positive (FP):{sum(FP_list)/len(FP_list)}")
        file.writelines(f"True Negative (TN):{sum(TN_list)/len(TN_list)}")
        file.writelines(f"False Negative (FN):{sum(FN_list)/len(FN_list)}")
        file.writelines(f'F1-score:{sum(F1_list)/len(F1_list)}')
        file.writelines(f"rank1={(sum(rank1_list)/3368)*100}%  rank5={(sum(rank5_list)/3368)*100}%  rank10={(sum(rank10_list)/3368)*100}%")

    print(f"총 소요 시간: {elapsed_time}초")
    print(f"필터링된 갤러리 이미지 개수: {sum(count_list)/len(count_list)}")
    print(f"필터링된 잔여 신원 개수: {sum(id_list)/len(id_list)}") #신원 개수 구하기
    print("True Positive (TP):", sum(TP_list)/len(TP_list))
    print("False Positive (FP):", sum(FP_list)/len(FP_list))
    print("True Negative (TN):", sum(TN_list)/len(TN_list))
    print("False Negative (FN):", sum(FN_list)/len(FN_list))
    print('F1-score:', sum(F1_list)/len(F1_list))
    print(f"rank1={(sum(rank1_list)/3368)*100}%  rank5={(sum(rank5_list)/3368)*100}%  rank10={(sum(rank10_list)/3368)*100}%")